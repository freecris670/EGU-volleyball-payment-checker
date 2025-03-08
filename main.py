from src.payment_checker import EventParser
from src.excluded_players import EXCLUDED_PLAYERS
import tkinter as tk
from tkinter import filedialog
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys
import re

# Словарь для персонализированных обращений
name_mappings = {
    "Maxim Зinovyev": "Макс",
    "Konstantin Podlesnyi": "Костя", 
    "Anya Vilkova": "Аня",
    "Kirill Kriukov": "Кирилл",
    "Ilya Kolomiytsev": "Илья",
    "Anna Kersilova": "Аня",
    "Kniazev Sergei": "Сергей",
    "Fedor Udalov": "Федя",
    "Sergey Konstantinov": "Сергей",
    "Ashot Gevorgyan": "Ашот",
    "Dmitry Kuznetsov": "Дмитрий",
    "Антирепрессант": "Араик",
    "Артем А": "Артем",
    "N": "Никита красивый",
    "Polina Matveeva": "Полина",
    "Daria Goltsova": "Дарья",
    "Viacheslav": "Слава",
    "Svetlana": "Светлана",
    "Valeriy": "Валерий",
    "Nikita Shabalin": "Никита",
    "Alice Koshechkina": "Алиса",
    "Matvey": "Матвей",
    "Ruslan": "Руслан",
    "Anatoliy Kudrin": "Толя",
    "Ivan": "Ваня",
    "Andrey Beloborodov": "Андрей",
    "Vanya": "Ваня",
    "Sergey Tkachuk": "Сергей",
    "Jek": "Женя",
    "Stas Maltsev": "Стас",
    "Noro Stambolyan": "Норо",
    "Narek Sargsyan": "Нарек",
    "Natali": "Натали",
    "Արարատ Մարտիրոսյան": "Арарат",
    "Artur Hovakanyan": "Артур",
    "Edgar": "Эдгар",
    "Greg": "Гриша",
    "Leonid": "Лёня",
    "Davo Abrahamyan": "Даво",
    "Dronte": "Андрей",
    "Lev Nikolsky": "Лев",
    "Vladimir": "Владимир",
    "Ivan Smirnov": "Иван",
    "Kristina Demirtshyan": "Кристина",
    "Ivan Nikolaev": "Ваня",
    # Добавьте другие соответствия имён по необходимости
}

def create_personal_message(name, dates):
    """Создание персонализированного сообщения"""
    # Персональное обращение (имя уже обработано без [N] в group_by_player)
    personal_name = name_mappings.get(name, name.split()[0])
    
    # Проверяем, есть ли игры с несколькими голосами
    multiple_votes_dates = []
    formatted_dates = []
    
    for date_info in dates:
        parts = date_info.split(' | ')
        date_display = parts[0]  # дата с днем недели
        votes = int(parts[-1]) if len(parts) > 2 else 1
        
        formatted_dates.append(date_display)
        
        # Учитываем только дополнительные голоса (больше 1)
        if votes > 1:
            additional_votes = votes - 1  # Вычитаем голос за себя
            multiple_votes_dates.append((date_display, additional_votes))
    
    # Форматируем даты для основного сообщения
    if len(formatted_dates) == 1:
        date_str = f"за игру в {formatted_dates[0]}"
    else:
        date_str = f"за игры в {', '.join(formatted_dates[:-1])} и {formatted_dates[-1]}"
    
    # Основное сообщение
    message = f"{personal_name}, привет!\n\nПодскажи, {date_str} картой оплачивал или наличкой?"
    
    # Если есть игры с дополнительными голосами, добавляем дополнительный вопрос
    if multiple_votes_dates:
        multiple_votes_message = "\n\nКстати, заметил, что у тебя "
        
        if len(multiple_votes_dates) == 1:
            date, additional_votes = multiple_votes_dates[0]
            vote_phrase = "дополнительный голос" if additional_votes == 1 else "дополнительных голоса"
            if additional_votes > 4:  # Для 5 и более используем "голосов"
                vote_phrase = "дополнительных голосов"
            multiple_votes_message += f"на игру в {date} есть {additional_votes} {vote_phrase}."
        else:
            vote_descriptions = []
            for date, additional_votes in multiple_votes_dates:
                vote_phrase = "дополнительный голос" if additional_votes == 1 else "дополнительных голоса"
                if additional_votes > 4:  # Для 5 и более используем "голосов"
                    vote_phrase = "дополнительных голосов"
                vote_descriptions.append(f"на игру в {date} - {additional_votes} {vote_phrase}")
            multiple_votes_message += f"есть несколько дополнительных голосов: {', '.join(vote_descriptions)}."
        
        multiple_votes_message += " Подскажи, эти дополнительные голоса тоже оплачены?"
        message += multiple_votes_message
    
    return message

def group_by_player(results):
    """Группировка дат игр по игрокам, учитывая разные игры в один день"""
    player_dates = defaultdict(list)
    
    # Структура для определения уникальных игр и подсчета голосов
    # формат: {(дата, время, базовое_имя): максимальное_количество_голосов}
    vote_tracking = {}
    
    # Словарь для хранения соответствия полных имен к базовым именам
    name_to_base_name = {}
    
    for result in results:
        # Разбираем результат, который включает и время игры
        parts = result.split(' | ')
        if len(parts) >= 2:
            date_info = parts[0]  # дата и день недели
            event_time = parts[1]  # время игры
            full_name = ' '.join(parts[2:])  # имя игрока
            
            # Проверяем, имеет ли имя пометку [N]
            vote_pattern = re.search(r'(.*?)\s*\[(\d+)\]$', full_name)
            votes = 1  # По умолчанию 1 голос
            
            if vote_pattern:
                base_name = vote_pattern.group(1).strip()
                votes = int(vote_pattern.group(2))
            else:
                base_name = full_name
            
            # Сохраняем связь полного имени с базовым (без [N])
            name_to_base_name[full_name] = base_name
            
            date_parts = date_info.split()
            date = date_parts[1] if len(date_parts) > 1 else date_parts[0]
            weekday = date_parts[0] if len(date_parts) > 1 else ""
            
            # Пропускаем исключённых игроков
            if base_name not in EXCLUDED_PLAYERS:
                # Формируем уникальный ключ для игры и базового имени игрока
                event_player_key = (date, event_time, base_name)
                
                # Обновляем количество голосов, выбирая максимальное значение
                current_votes = vote_tracking.get(event_player_key, 0)
                vote_tracking[event_player_key] = max(current_votes, votes)
                
                # Формируем отображаемую дату с временем
                display_date = f"{weekday} {date} | {event_time}"
                
                # Добавляем дату, если это первое появление базового имени
                if base_name not in [name_to_base_name.get(p) for p in player_dates.keys()]:
                    player_dates[base_name].append(display_date)
                # Если базовое имя уже есть, проверяем наличие этой даты
                elif display_date not in player_dates[base_name]:
                    player_dates[base_name].append(display_date)
    
    # Обогащаем информацию о датах количеством голосов
    enriched_player_dates = {}
    for player, dates in player_dates.items():
        enriched_dates = []
        for date in dates:
            date_parts = date.split(' | ')
            date_info = date_parts[0]
            date_parts_split = date_info.split()
            date_only = date_parts_split[1] if len(date_parts_split) > 1 else date_parts_split[0]
            event_time = date_parts[1] if len(date_parts) > 1 else ""
            
            event_player_key = (date_only, event_time, player)
            votes = vote_tracking.get(event_player_key, 1)
            
            # Формат: "дата | время | количество_голосов"
            enriched_date = f"{date} | {votes}"
            enriched_dates.append(enriched_date)
            
        enriched_player_dates[player] = enriched_dates
    
    return enriched_player_dates

def create_excel_report(player_dates):
    """Создание Excel-отчета с результатами"""
    wb = Workbook()
    
    # Создаем первую вкладку с оплатами
    ws_payments = wb.active
    ws_payments.title = "Оплаты"

    # Заголовки столбцов
    headers = ["Дата", "Время", "Имя", "Реальное имя", "Голоса", "Вид оплаты", "Статус"]
    for col, header in enumerate(headers, 1):
        ws_payments.cell(row=1, column=col, value=header)

    # Создаем выпадающие списки
    payment_type_dv = DataValidation(
        type="list",
        formula1='"картой,наличкой"',
        allow_blank=True
    )
    
    status_dv = DataValidation(
        type="list",
        formula1='"Запрос оплаты,Оплачено,Отсутствовал"',
        allow_blank=True
    )

    ws_payments.add_data_validation(payment_type_dv)
    ws_payments.add_data_validation(status_dv)

    # Словарь для отслеживания общего количества голосов на каждую игру
    event_total_votes = {}

    # Собираем данные для сортировки
    payment_data = []
    for player, dates in player_dates.items():
        # Получаем персональное обращение (имя уже без [N])
        personal_name = name_mappings.get(player, player.split()[0])
        
        for date_info in dates:
            # Разделяем дату и время
            parts = date_info.split(' | ')
            display_date = parts[0]
            event_time = parts[1] if len(parts) > 1 else ""
            votes = int(parts[-1]) if len(parts) > 2 else 1
            
            # Разбиваем дату на части для сортировки
            day_month = display_date.split()[-1]  # получаем только "DD.MM"
            day, month = day_month.split('.')
            
            # Формируем дату для сортировки (месяц.день)
            sort_date = f"{month}.{day}"
            
            # Ключ для отслеживания уникальной игры для данного игрока
            event_key = (display_date, event_time, player)
            
            # Сохраняем информацию о количестве голосов
            event_total_votes[event_key] = votes
            
            # Создаем отдельную запись для КАЖДОГО голоса
            for vote_index in range(votes):
                payment_data.append({
                    'sort_date': sort_date,
                    'display_date': display_date,
                    'event_time': event_time,
                    'player': player,
                    'personal_name': personal_name,
                    'votes': votes,
                    'event_key': event_key,
                    'day': int(day),
                    'month': int(month)
                })
    
    # Сортируем данные по дате
    payment_data.sort(key=lambda x: (x['month'], x['day']))
    
    # Получаем первую и последнюю даты из отсортированного списка
    start_date = None
    end_date = None
    
    if payment_data:
        first_item = payment_data[0]
        last_item = payment_data[-1]
        
        # Форматируем даты в виде "DD.MM"
        start_date = f"{first_item['day']:02d}.{first_item['month']:02d}"
        end_date = f"{last_item['day']:02d}.{last_item['month']:02d}"
    
    # Заполняем данные на первой вкладке в отсортированном порядке
    row = 2
    for entry in payment_data:
        ws_payments.cell(row=row, column=1, value=entry['display_date'])
        ws_payments.cell(row=row, column=2, value=entry['event_time'])
        ws_payments.cell(row=row, column=3, value=entry['player'])  # Оригинальное имя
        ws_payments.cell(row=row, column=4, value=entry['personal_name'])  # Персонализированное имя
        ws_payments.cell(row=row, column=5, value=entry['votes'])  # Общее количество голосов
        
        # Применяем валидацию к ячейкам
        payment_type_dv.add(f'F{row}')
        status_dv.add(f'G{row}')
        
        row += 1

    # Автоматическая настройка ширины столбцов
    for column in ws_payments.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_payments.column_dimensions[column_letter].width = adjusted_width

    # Создаем вторую вкладку с личными сообщениями
    ws_messages = wb.create_sheet(title="Личные сообщения")
    ws_messages.column_dimensions['A'].width = 30
    ws_messages.column_dimensions['B'].width = 70

    # Заголовки для личных сообщений
    ws_messages.cell(row=1, column=1, value="Имя")
    ws_messages.cell(row=1, column=2, value="Сообщение")

    # Заполняем личные сообщения
    row = 2
    for player, dates in player_dates.items():
        message = create_personal_message(player, dates)
        ws_messages.cell(row=row, column=1, value=player)
        ws_messages.cell(row=row, column=2, value=message)
        row += 1

    # Изменяем логику сохранения файла - без диалога с пользователем
    try:
        # Получаем путь к директории, где находится исполняемый файл
        if getattr(sys, 'frozen', False):
            # Если это скомпилированный exe
            application_path = os.path.dirname(sys.executable)
        else:
            # Если это python скрипт
            application_path = os.path.dirname(os.path.abspath(__file__))
            
        # Создаем директорию для отчетов, если её нет
        reports_dir = os.path.join(application_path, 'Оплаты за волейбол')
        if not os.path.exists(reports_dir):
            os.makedirs(reports_dir)
        
        # Формируем имя файла
        filename = "Оплаты за волейбол"
        if start_date and end_date:
            if start_date == end_date:
                filename += f" {start_date}"
            else:
                filename += f" {start_date}-{end_date}"
        filename += ".xlsx"
        
        # Формируем полный путь к файлу
        file_path = os.path.join(reports_dir, filename)
        
        # Сохраняем файл без запроса пользователя
        wb.save(file_path)
        print(f"\nОтчет автоматически сохранен в файл: {file_path}")
            
    except Exception as e:
        print(f"\nОшибка при сохранении файла: {e}")

def main():
    # Создаём корневое окно и сразу скрываем его
    root = tk.Tk()
    root.withdraw()

    # Открываем диалоговое окно выбора файла
    print("Выберите текстовый файл с данными события...")
    file_path = filedialog.askopenfilename(
        title="Выберите файл с данными события",
        filetypes=[("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")]
    )

    if not file_path:
        print("Файл не выбран. Программа завершена.")
        return

    try:
        # Читаем содержимое файла
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()

        # Создаём парсер и обрабатываем текст
        parser = EventParser()
        results = parser.parse_event_text(text)
        
        # Группируем результаты по игрокам
        player_dates = group_by_player(results)
        
        # Выводим результаты и создаём личные сообщения
        print("\nРезультаты обработки и личные сообщения:")
        print("=" * 50)
        for player, dates in player_dates.items():
            print(f"\nИгрок: {player}")
            print(f"Даты игр: {', '.join(dates)}")
            message = create_personal_message(player, dates)
            print("\nЛичное сообщение:")
            print(message)
            print("-" * 30)

        # Создаем Excel-отчет
        create_excel_report(player_dates)

    except FileNotFoundError:
        print("Ошибка: Файл не найден.")
    except Exception as e:
        print(f"Произошла ошибка при чтении файла: {e}")

if __name__ == "__main__":
    main() 